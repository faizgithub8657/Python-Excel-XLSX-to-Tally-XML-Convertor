"""
converter.py — ABHL & Associates Excel → Tally Prime XML converter
Core conversion logic, importable by both the CLI script and the Streamlit web app.

KEY FEATURES
============
- Auto-detects header row by scanning for known column names
  (so the user's data can start at A1, C3, or anywhere else)
- Picks the right worksheet by name OR auto-detects the sheet that has
  voucher-style headers
- Handles dates in many formats (datetime cells, DD-MM-YYYY, DD/MM/YYYY,
  YYYY-MM-DD, etc.) and outputs YYYYMMDD as Tally requires
- Tally sign convention:
    ISDEEMEDPOSITIVE = YES  -> debit leg  (AMOUNT negative)
    ISDEEMEDPOSITIVE = NO   -> credit leg (AMOUNT positive)
- For PAYMENT  : Dr columns = accounting Dr (party), Cr columns = bank
  For RECEIPT  : swapped — party in "Dr_Ledger" column becomes the Cr leg
                  in the XML; bank in "Cr_Ledger" column becomes Dr leg
  For JOURNAL/CONTRA/others: straight Dr=Dr, Cr=Cr
- Skips blank rows automatically
- Returns both the XML string AND a row-level log so the UI can show
  exactly which rows succeeded, which were skipped, and why

USAGE (as library)
==================
    from converter import convert_workbook
    xml_text, log = convert_workbook(file_like_or_path, sheet_name="Sheet1")

Reference: Tally Solutions, Tally Prime — Import & Export of Data via XML
           (help.tallysolutions.com)
"""

from __future__ import annotations
import io
from datetime import datetime, date
from xml.sax.saxutils import escape
from typing import Tuple, List, Dict, Any

import openpyxl


# Columns we expect (case-insensitive match). Auto-detection scans every
# cell of every row in the chosen sheet until it finds a row containing
# at least this many of these names — that row becomes the header.
EXPECTED_HEADERS = {
    "DATE", "VCHTYPE", "DR_LEDGER_1", "CR_LEDGER_1", "DR_AMT_1", "CR_AMT_1"
}
MIN_HEADER_MATCHES = 4  # row must contain at least 4 expected names to qualify

LEG_PAIRS = 4  # supports Dr_Ledger_1..4 and Cr_Ledger_1..4


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------

def _fmt_date(val) -> str:
    """Convert any reasonable date input into Tally's YYYYMMDD format."""
    if val is None or val == "":
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y%m%d")
    s = str(val).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y",
                "%d %b %Y", "%d-%B-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    return s  # caller will flag invalid dates downstream


def _clean(val) -> str:
    if val is None:
        return ""
    return escape(str(val).strip())


def _num(val):
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _fmt_amt(amt: float) -> str:
    """Match the reference file — integers stay integers, fractions get 2 dp."""
    if float(amt).is_integer():
        return str(int(amt))
    return f"{amt:.2f}"


# --------------------------------------------------------------------------
# sheet + header detection
# --------------------------------------------------------------------------

def _pick_worksheet(wb, preferred_name: str | None):
    """Return the worksheet most likely to hold voucher data."""
    if preferred_name and preferred_name in wb.sheetnames:
        return wb[preferred_name]
    # fall back: scan every sheet and return the first one that contains
    # the expected voucher headers
    for ws in wb.worksheets:
        if _find_header_row(ws)[0] is not None:
            return ws
    # last resort
    return wb.active


def _find_header_row(ws) -> Tuple[int | None, int | None, Dict[str, int]]:
    """
    Scan the worksheet for a row that contains the expected voucher headers.
    Returns (header_row_number, header_start_col, {normalised_name -> col_idx}).
    Row & col numbers are 1-indexed (openpyxl style).
    Returns (None, None, {}) if no header row is found.
    """
    max_scan_rows = min(ws.max_row or 0, 50)  # only scan top 50 rows
    for r in range(1, max_scan_rows + 1):
        row_cells = list(ws[r])
        # build name -> col_idx map for this row
        hdr_map = {}
        for cell in row_cells:
            if cell.value is None:
                continue
            name = str(cell.value).strip().upper().replace(" ", "_")
            if name:
                hdr_map[name] = cell.column  # 1-indexed
        matches = len(EXPECTED_HEADERS & set(hdr_map.keys()))
        if matches >= MIN_HEADER_MATCHES:
            start_col = min(hdr_map.values())
            return r, start_col, hdr_map
    return None, None, {}


# --------------------------------------------------------------------------
# voucher XML builder
# --------------------------------------------------------------------------

def _build_voucher_xml(row_data: Dict[str, Any]) -> Tuple[str | None, str]:
    """
    Build one <TALLYMESSAGE> block from one logical row.
    Returns (xml_or_None, reason_if_skipped).
    """
    vch_type = (row_data.get("VCHTYPE") or "").strip().upper()
    if not vch_type:
        return None, "no VCHTYPE"

    date_raw = row_data.get("DATE")
    date_str = _fmt_date(date_raw)
    if not date_str or len(date_str) != 8 or not date_str.isdigit():
        return None, f"invalid date: {date_raw!r}"

    vch_num = _clean(row_data.get("VCH_NUM"))
    narration = _clean(row_data.get("NARRATION"))

    # collect legs
    dr_legs, cr_legs = [], []
    for i in range(1, LEG_PAIRS + 1):
        dl = row_data.get(f"DR_LEDGER_{i}")
        da = _num(row_data.get(f"DR_AMT_{i}"))
        cl = row_data.get(f"CR_LEDGER_{i}")
        ca = _num(row_data.get(f"CR_AMT_{i}"))
        if dl and da and da > 0:
            dr_legs.append((str(dl).strip(), da))
        if cl and ca and ca > 0:
            cr_legs.append((str(cl).strip(), ca))

    if not dr_legs or not cr_legs:
        return None, "missing Dr or Cr leg"

    # balance check
    dr_total = sum(a for _, a in dr_legs)
    cr_total = sum(a for _, a in cr_legs)
    if abs(dr_total - cr_total) > 0.01:
        return None, f"Dr {dr_total} <> Cr {cr_total} (not balanced)"

    # roles based on voucher type (see module docstring)
    if vch_type == "RECEIPT":
        accounting_dr = cr_legs  # bank
        accounting_cr = dr_legs  # party
        party = dr_legs[0][0]
        first_block, second_block = accounting_cr, accounting_dr
        first_isdp, second_isdp = "NO", "YES"
        first_sign, second_sign = "", "-"
    else:
        accounting_dr = dr_legs
        accounting_cr = cr_legs
        party = dr_legs[0][0]
        first_block, second_block = accounting_dr, accounting_cr
        first_isdp, second_isdp = "YES", "NO"
        first_sign, second_sign = "-", ""

    # build ledger entry blocks
    entries: List[str] = []
    for name, amt in first_block:
        entries.append(
            "\t\t\t\t\t\t<ALLLEDGERENTRIES.LIST>\n"
            f"\t\t\t\t\t\t\t<LEDGERNAME>{escape(name)}</LEDGERNAME>\n"
            "\t\t\t\t\t\t\t<REMOVEZEROENTRIES>NO</REMOVEZEROENTRIES>\n"
            "\t\t\t\t\t\t\t<LEDGERFROMITEM>NO</LEDGERFROMITEM>\n"
            f"\t\t\t\t\t\t\t<ISDEEMEDPOSITIVE>{first_isdp}</ISDEEMEDPOSITIVE>\n"
            f"\t\t\t\t\t\t\t<AMOUNT>{first_sign}{_fmt_amt(amt)}</AMOUNT>\n"
            "\t\t\t\t\t\t</ALLLEDGERENTRIES.LIST>"
        )
    for name, amt in second_block:
        entries.append(
            "\t\t\t\t\t\t<ALLLEDGERENTRIES.LIST>\n"
            f"\t\t\t\t\t\t\t<LEDGERNAME>{escape(name)}</LEDGERNAME>\n"
            "\t\t\t\t\t\t\t<REMOVEZEROENTRIES>NO</REMOVEZEROENTRIES>\n"
            "\t\t\t\t\t\t\t<LEDGERFROMITEM>NO</LEDGERFROMITEM>\n"
            f"\t\t\t\t\t\t\t<ISDEEMEDPOSITIVE>{second_isdp}</ISDEEMEDPOSITIVE>\n"
            f"\t\t\t\t\t\t\t<AMOUNT>{second_sign}{_fmt_amt(amt)}</AMOUNT>\n"
            "\t\t\t\t\t\t</ALLLEDGERENTRIES.LIST>"
        )

    voucher_xml = (
        "\t\t\t\t<TALLYMESSAGE xmlns:UDF=\"TallyUDF\">\n"
        f"\t\t\t\t\t<VOUCHER ACTION=\"Create\" VCHTYPE=\"{escape(vch_type)}\">\n"
        f"\t\t\t\t\t\t<VOUCHERTYPENAME>{escape(vch_type)}</VOUCHERTYPENAME>\n"
        f"\t\t\t\t\t\t<DATE>{date_str}</DATE>\n"
        f"\t\t\t\t\t\t<VOUCHERNUMBER>{vch_num}</VOUCHERNUMBER>\n"
        f"\t\t\t\t\t\t<PARTYLEDGERNAME>{escape(party)}</PARTYLEDGERNAME>\n"
        f"\t\t\t\t\t\t<NARRATION>{narration}</NARRATION>\n"
        f"\t\t\t\t\t\t<EFFECTIVEDATE>{date_str}</EFFECTIVEDATE>\n"
        + "\n".join(entries) +
        "\n\t\t\t\t\t</VOUCHER>\n"
        "\t\t\t\t</TALLYMESSAGE>"
    )
    return voucher_xml, ""


# --------------------------------------------------------------------------
# public entry point
# --------------------------------------------------------------------------

def convert_workbook(file_input, sheet_name: str | None = "Sheet1"
                     ) -> Tuple[str, List[Dict[str, Any]]]:
    """
    Convert an Excel workbook into Tally XML.

    Parameters
    ----------
    file_input : str | bytes | BytesIO | UploadedFile
        Path to the .xlsx file OR a file-like object (Streamlit uploaded file).
    sheet_name : str | None
        Preferred sheet name; falls back to auto-detection.

    Returns
    -------
    xml_text : str   The full <ENVELOPE>...</ENVELOPE> XML.
    log      : list  Row-level log entries with keys:
                       row, status ('ok' | 'skipped'), reason, voucher_num
    """
    if isinstance(file_input, (bytes, bytearray)):
        file_input = io.BytesIO(file_input)
    wb = openpyxl.load_workbook(file_input, data_only=True)
    ws = _pick_worksheet(wb, sheet_name)

    header_row, _, hdr_map = _find_header_row(ws)
    if header_row is None:
        raise ValueError(
            f"No header row found in sheet '{ws.title}'. Expected to find a "
            f"row containing at least {MIN_HEADER_MATCHES} of these column "
            f"names: {sorted(EXPECTED_HEADERS)}"
        )

    vouchers: List[str] = []
    log: List[Dict[str, Any]] = []

    # iterate data rows
    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        # build {NORMALISED_HEADER: value} for this row
        row_data: Dict[str, Any] = {}
        for name, col in hdr_map.items():
            row_data[name] = ws.cell(row=r, column=col).value

        # skip totally blank rows silently
        if not any(v not in (None, "") for v in row_data.values()):
            continue

        xml, reason = _build_voucher_xml(row_data)
        if xml:
            vouchers.append(xml)
            log.append({"row": r, "status": "ok", "reason": "",
                        "voucher_num": _clean(row_data.get("VCH_NUM"))})
        else:
            log.append({"row": r, "status": "skipped", "reason": reason,
                        "voucher_num": _clean(row_data.get("VCH_NUM"))})

    envelope = (
        "<ENVELOPE>\n"
        "\t<HEADER>\n"
        "\t\t<TALLYREQUEST>Import Data</TALLYREQUEST>\n"
        "\t</HEADER>\n"
        "\t<BODY>\n"
        "\t\t<IMPORTDATA>\n"
        "\t\t\t<REQUESTDESC>\n"
        "\t\t\t\t<REPORTNAME>All Masters</REPORTNAME>\n"
        "\t\t\t\t<STATICVARIABLES>\n"
        "\t\t\t\t\t<SVCURRENTCOMPANY></SVCURRENTCOMPANY>\n"
        "\t\t\t\t</STATICVARIABLES>\n"
        "\t\t\t</REQUESTDESC>\n"
        "\t\t\t<REQUESTDATA>\n"
        + "\n".join(vouchers) +
        "\n\t\t\t</REQUESTDATA>\n"
        "\t\t</IMPORTDATA>\n"
        "\t</BODY>\n"
        "</ENVELOPE>\n"
    )
    return envelope, log


# --------------------------------------------------------------------------
# CLI usage:  python converter.py input.xlsx output.xml [sheet_name]
# --------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Usage: python converter.py input.xlsx output.xml [sheet_name]")
        sys.exit(1)
    src, dst = sys.argv[1], sys.argv[2]
    sheet = sys.argv[3] if len(sys.argv) > 3 else "Sheet1"
    xml, log = convert_workbook(src, sheet)
    with open(dst, "w", encoding="utf-8") as f:
        f.write(xml)
    ok = sum(1 for e in log if e["status"] == "ok")
    skip = sum(1 for e in log if e["status"] == "skipped")
    print(f"Wrote {ok} vouchers to {dst}  (skipped {skip} rows)")
    for e in log:
        if e["status"] == "skipped":
            print(f"  - row {e['row']} ({e['voucher_num']}): {e['reason']}")
